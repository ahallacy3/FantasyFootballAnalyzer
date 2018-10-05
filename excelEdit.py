from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScale, FormatObject, ColorScaleRule
import string
from openpyxl.styles import Color, Font


class excelEdit:
    def reader(self, num_teams, num_weeks, fileName, sheetName):
        wb = load_workbook(fileName)
        dataSheet = wb[sheetName]

        num_rows = 3 + (num_teams / 2) * num_weeks + 1
        matchData = []
        weekData = []
        seasonData = []
        for row in dataSheet['A3:D'+str(int(num_rows))]:
            for cell in row:
                matchData.append(cell.value)
            if matchData[1] is not None:
                weekData.append(matchData)
                matchData = []
            else:
                seasonData.append(weekData)
                matchData = []
                weekData = []
        return seasonData

    def writer(self, results, resultsFileName):
        wb = Workbook()
        sheet = wb['Sheet']
        for col in range(1, len(results) + 1):
            for row in range(1, len(results[col - 1]) + 1):
                sheet.cell(column=col, row=row, value=results[col - 1][row - 1])

        self.conditionalFormatter(sheet, results)
        self.headerFonter(sheet, results)
        self.columnWidth(sheet, results)
        wb.save(resultsFileName)

    def conditionalFormatter(self, sheet, results):
        rule = ColorScaleRule(start_type='percentile', start_value=0, start_color='F8696B',
                              mid_type = 'percentile', mid_value = 50, mid_color = 'FFEB84',
                              end_type = 'percentile', end_value = 100, end_color = '63BE7B')
        for i in range(1, len(results)):
            cellRange = string.ascii_uppercase[i] + str(2) + ':' + string.ascii_uppercase[i] + str(len(results[i]) - 2)
            sheet.conditional_formatting.add(cellRange, rule)

    def headerFonter(self, sheet, results):
        font = Font(bold = True, size = 10)
        for i in range (1, len(results)):
            sheet[string.ascii_uppercase[i] + str(1)].font = font

    def columnWidth(self, sheet, results):
        widthMap = {'Total': 4.72}
        widthMap['Team Name'] = 15.5
        widthMap['Average'] = 7.06
        widthMap['Median'] = 6.39
        widthMap['Range'] = 5.72
        widthMap['STDEV'] = 5.72
        widthMap['Real Wins'] = 8.48
        widthMap['Wins Against Everyone'] = 18.61
        widthMap['Normalized WAE'] = 13.84
        widthMap['Win Diff'] = 7.06
        widthMap['SoS'] = 3.5
        widthMap['Normalized Losses'] = 15.28
        widthMap['Loss Diff'] = 7.28
        widthMap['Expected Wins So Far'] = 17.72
        widthMap['How bad has your schedule fucked you?'] = 13.72

        for column_cells in sheet.columns:
            length = widthMap.get(column_cells[0].value, 1)
            sheet.column_dimensions[column_cells[0].column].width = length